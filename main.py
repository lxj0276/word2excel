# -*- coding: utf-8 -*-
"""
本脚本实现从input文件夹的docx文件的相关数据
到output文件夹的excel表格


可提取的word文件示例：
本委托人于2015年12月24日提取第【201】期委托资产现金资产价值人民币2,630,136.99元（大写人民币贰佰陆拾叁万零壹佰叁拾陆元玖角玖分），其中委托资产本金人民币0.00元，委托资产收益人民币2,630,136.99元，按照定向资产管理合同的规定，管理人签收后回传委托人并与其确认收到。
提取其中的日期、期数、金额信息
"""
import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')  # 编译环境utf8
from glob import glob
import re
import time
try:
    from docx import Document
except ImportError:
    print(u'缺少模块python-docx，正在自动安装')
    import subprocess
    p = subprocess.Popen('easy_install python-docx', shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    print(p.stdout.readlines())
    for line in p.stdout.readlines():
        print(line)
    retval = p.wait()
    from docx import Document
    # raise
try:
    import openpyxl
except ImportError:
    print(u'缺少模块openpyxl，正在自动安装')
    import subprocess
    p = subprocess.Popen('easy_install openpyxl', shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    print(p.stdout.readlines())
    for line in p.stdout.readlines():
        print(line)
    retval = p.wait()
    import openpyxl
    # raise
##################################这是彩色打印
import ctypes
STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE= -11
STD_ERROR_HANDLE = -12


FOREGROUND_BLACK = 0x0
FOREGROUND_BLUE = 0x01 # text color contains blue.
FOREGROUND_GREEN= 0x02 # text color contains green.
FOREGROUND_RED = 0x04 # text color contains red.
FOREGROUND_INTENSITY = 0x08 # text color is intensified.

BACKGROUND_BLUE = 0x10 # background color contains blue.
BACKGROUND_GREEN= 0x20 # background color contains green.
BACKGROUND_RED = 0x40 # background color contains red.
BACKGROUND_INTENSITY = 0x80 # background color is intensified.
#上面这一大段都是在设置前景色和背景色，其实可以用数字直接设置，我的代码直接用数字设置颜色


class Color:
    std_out_handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)

    def set_cmd_color(self, color, handle=std_out_handle):
        bool = ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)
        return bool

    def reset_color(self):
        self.set_cmd_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE | FOREGROUND_INTENSITY)
        #初始化颜色为黑色背景，纯白色字，CMD默认是灰色字体的

    def print_red_text(self, print_text):
        self.set_cmd_color(4 | 8)
        print(print_text)
        self.reset_color()
        #红色字体

    def print_green_text(self, print_text):
        self.set_cmd_color(FOREGROUND_GREEN | FOREGROUND_INTENSITY)
        # c = raw_input(print_text.encode('gbk'))
        # c = raw_input(print_text)
        print(print_text)
        self.reset_color()
        # return c

    def print_yellow_text(self, print_text):
        self.set_cmd_color(6 | 8)
        print(print_text)
        self.reset_color()
        #黄色字体

    def print_blue_text(self, print_text):
        self.set_cmd_color(1 | 10)
        print(print_text)
        self.reset_color()
        #蓝色字体


clr = Color()
clr.set_cmd_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE | FOREGROUND_INTENSITY)
# clr.print_red_text('red')
# clr.print_green_text("green")
# clr.print_blue_text('blue')
# clr.print_yellow_text('yellow')
##########################################


PROJECT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))


re_pats = {
    'date': u'\d+年\d+月\d+日',
    'stage': u'【(\d+)】',
    'money': u'人民币(.*?)元',
}


def get_filename_by_path(path, forbid_word=''):
    searched_filenames = glob(path)
    return [i for i in searched_filenames if forbid_word not in i ]

def parse_text_by_repat(text, re_pat):
    find = re.findall(re_pat, text)
    return (find[0].strip() if find else None)


def read_docx(filename):
    result = {
        'date': None,
        'stage': None,
        'money': None,
        # 'money1': None,
    }
    document = Document(filename)
    locate_text = u'本委托人'  # 定位段落的关键字
    max_search_para = 10  # 最多搜索前n段来定位段落
    print(u'定位并解析数据')
    for para_num, para in enumerate(document.paragraphs):
        if para_num < max_search_para and locate_text in para.text:
            # result1 = parse_text_by_repat(para.text, u'人民币(.*?)元')

            for key in result.keys():
                result[key] = parse_text_by_repat(para.text, re_pats[key])

            # # 可以更改
            # result2 = re.findall(u'人民币(.*?)元', para.text)
            # if len(result2) >= 2:
            #     result['money1'] = result2[1]
            # # 停止更改
            return result
            break  # 包含关键字则退出循环搜索
        elif para_num >= max_search_para:
            raise IOError(u'Error！文件前%s段不包含指定的文字：%s，请检查！' % (max_search_para, locate_text))

def write_excel(excel_name, result_dicts):
    from openpyxl.workbook import Workbook

    #ExcelWriter,里面封装好了对Excel的写操作
    from openpyxl.writer.excel import ExcelWriter

    #get_column_letter函数将数字转换为相应的字母，如1-->A,2-->B
    from openpyxl.cell import get_column_letter

    from openpyxl.reader.excel import load_workbook

    if os.path.isfile(excel_name):
        # #读取excel2007文件
        wb = load_workbook(excel_name)
    else:
        #新建一个workbook
        wb = Workbook()


    #新建一个excelWriter
    ew = ExcelWriter(workbook = wb)

    #设置文件输出路径与名称
    dest_filename = excel_name

    # # 获取第一个sheet
    try:
        ws = wb.get_sheet_by_name('sheet1')
    except KeyError:
        ws = wb.worksheets[0]
        ws.title = "sheet1"


    #第一个sheet是ws
    # ws = wb.worksheets[0]

    # #设置ws的名称
    # ws.title = "sheet1"

    line = 1
    print(u'定位写入坐标')
    while ws.cell("A%s" % line).value:
        # print(ws.cell("A%s" % line).value)
        line += 1
    print(u'从第%s行开始写入' % line)

    if not os.path.isfile(excel_name):
        ws.cell("A%s" % line).value=u'期数'
        ws.cell("B%s" % line).value=u'定向计划提取'
        ws.cell("C%s" % line).value=u'大写'
        ws.cell("D%s" % line).value=u'到期时间'
        ws.cell("E%s" % line).value=u'交付金额'
        line += 1
    for i, result in enumerate(result_dicts):
        print(u'正在写入第%s条数据到excel' % (i+1))
        ws.cell("A%s" % line).value=result['stage']
        ws.cell("B%s" % line).value=result['money']
        ws.cell("C%s" % line).value=''
        ws.cell("D%s" % line).value=result['date']
        ws.cell("E%s" % line).value=result['money']
        line += 1

    #最后保存文件
    ew.save(filename=excel_name)

def main():
    print(u'开始执行')
    print(u'从input文件夹查找docx文件')
    filenames = get_filename_by_path('input/*.docx', '~$')
    result_dicts = []
    for filename in filenames:
        print(u'读取文件：')
        clr.print_blue_text(os.path.basename(filename))
        # print(os.path.basename(filename))
        result_dicts.append(read_docx(filename))
    save_filename = 'output/output.xlsx'
    # save_filename = 'output/output%s.xlsx' % int(time.time())
    write_excel(save_filename, result_dicts)
    print(u'执行完毕，文件保存至')
    clr.print_blue_text(save_filename)
    # print(save_filename)
    print(u'敲击回车结束运行')
    raw_input()

if __name__ == '__main__':
    main()
